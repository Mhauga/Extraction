import requests
import openpyxl

# Open a workbook and the active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Farmers Markets"

def get_business_lists(city, last_row):
    print(city)

    # API Definition
    # my_API_Key = "insert Yelp API key"
    endPoint = "https://api.yelp.com/v3/businesses/search"
    api_headers = {'Authorization': 'bearer {}'.format(my_API_Key)}

    # Parameters
    offset = 0
    limit = 5
    total = 1000

    # Variable declarations
    biz_name = []
    biz_address = []
    biz_phone = []
    biz_reviews = []

    while offset < total:
        parameters = {"term": "Farmers Markets",
                      "limit": limit,
                     "location": city,
                     "sort_by": "rating",
                     "offset": offset}

        #  Make the request and then convert the json to a dictionary
        businesses_json_response = requests.get(url=endPoint, params=parameters, headers=api_headers)
        businesses = businesses_json_response.json()

        # Update Total
        try:
            total = businesses["total"]
            #  print(len(businesses["businesses"]))
        except:
            print("Total is less than {}".format(limit))
            break

        # Append business names, addresses, and phones to their respective lists
        for business in businesses["businesses"]:
            biz_name.append(business["name"])
            biz_address.append(business["location"]["address1"])
            biz_phone.append(business['display_phone'])
            biz_reviews.append(business['review_count'])

        # Write data of excel function
        last_row = print_to_excel(biz_name, biz_address, biz_phone, biz_reviews, city,last_row)

        #  print("Total = {}; Limit = {}; Offset = {}".format(total, limit, offset))

        if total < limit:
            limit = total
            offset += limit
        elif total >= limit:
            offset += limit

    return last_row


def print_to_excel(names, addresses, phones, reviews, city, last_row):
    # Write data to excel
    elem = 0
    while elem < len(names):
        ws.cell(row = last_row + elem + 1, column = 1).value = names[elem]
        ws.cell(row = last_row + elem + 1, column = 2).value = addresses[elem]
        ws.cell(row = last_row + elem + 1, column = 3).value = phones[elem]
        ws.cell(row = last_row + elem + 1, column = 4).value = city
        ws.cell(row = last_row + elem + 1, column = 5).value = "ID"
        ws.cell(row = last_row + elem + 1, column = 6).value = reviews[elem]
        elem += 1

    last_row = ws.max_row

    return last_row


def main():

    MSA_Columbus = (
        "Columbus", "Dublin", "Newark", "Delaware", "Lancaster", "Pickerington", "London", "Marysville", "Circleville",
        "Marion", "Zanesville", "Chillicothe", "New Lexington", "Cambridge", "Washington Court House")
    MSA_Dayton = ("Centerville", "Dayton", "Kettering", "Beavercreek", "Huber Heights", "Fairborn", "Miamisburg",
                  "West Carrollton", "Springfield", "Urbana", "Greenville", "Sidney")
    MSA_Indianapolis = ("Indianapolis", "Carmel", "Fishers", "Noblesville", "Greenwood", "Anderson", "Lawrence",
                        "Westfield", "Plainfield", "Zionsville", "Brownsburg", "Franklin", "Greenfrield", "Shelbyville",
                        "Avon", "Lebanon", "Beech Grove", "Speedway", "Martinsville",
                        "Greencastle","Danville", "Moorseville")

    last_row = 0 # Keeps track of position in Excel
    for city in MSA_Indianapolis:
        last_row = get_business_lists(city, last_row)


# BEGIN PROGRAM
main()
wb.save("MSA_Indianapolis_Farmers_Markets.xlsx")
